import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import (
    BM, TF, BMS, DEC, DECX, STATS, MIN_INT, TIMEFRAMES,
    INTERVALS_JSON, OUTPUTS_DIR, CHECKPOINT_DIR, XL_MET, XL_STAT,
    DT_XL_FORMAT, DT_XL_FILENAME, ENTROPIC_INTERVALS, SCRIPTS_DIR,
    RR_CSV, XL_BASE,
)

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import glob
import json
import warnings
import traceback
import argparse
import subprocess
import multiprocessing as mp
import antropy as ant
from scipy import signal
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import nolds
import neurokit2 as nk
import shutil

warnings.filterwarnings('ignore')


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
''' Interval Ledger System                                                            '''
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class IntervalLedger:
    """
    Stores processed coverage as merged, non-overlapping [beg_ms, end_ms] intervals.
    File format (JSON):
    {
      "schema": 1,
      "intervals": [[beg_ms, end_ms], ...]  # inclusive start, inclusive end
    }
    """
    def __init__(self, path: str):
        self.path = path
        self.schema = 1
        self.intervals = []
        self._loaded = False

    @staticmethod
    def _merge(intervals):
        if not intervals:
            return []
        intervals = sorted(intervals, key=lambda x: (x[0], x[1]))
        merged = [intervals[0][:]]
        for s, e in intervals[1:]:
            if s <= merged[-1][1] + 1:
                merged[-1][1] = max(merged[-1][1], e)
            else:
                merged.append([s, e])
        return merged

    @staticmethod
    def _subtract(base_interval, covered):
        S, E = base_interval
        gaps = []
        cur = S
        for s, e in covered:
            if e < S:
                continue
            if s > E:
                break
            if s > cur:
                gaps.append([cur, min(E, s - 1)])
            cur = max(cur, e + 1)
            if cur > E:
                break
        if cur <= E:
            gaps.append([cur, E])
        return gaps

    def load(self):
        if self._loaded:
            return
        try:
            if os.path.exists(self.path):
                with open(self.path, "r") as f:
                    data = json.load(f)
                if isinstance(data, dict) and "intervals" in data:
                    self.intervals = self._merge([[int(s), int(e)] for s, e in data["intervals"]])
                else:
                    self.intervals = []
            else:
                self.intervals = []
        except Exception as e:
            print(f"[WARNING] ❗ Could not load interval ledger at '{self.path}': {e}. Starting empty.")
            self.intervals = []
        self._loaded = True

    def save(self):
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        payload = {"schema": self.schema, "intervals": self._merge(self.intervals)}
        with open(self.path, "w") as f:
            json.dump(payload, f, indent=2)

    def add_intervals(self, new_intervals_ms):
        self.load()
        cleaned = []
        for s, e in new_intervals_ms:
            if s is None or e is None:
                continue
            s, e = int(s), int(e)
            if e < s:
                s, e = e, s
            cleaned.append([s, e])
        self.intervals = self._merge(self.intervals + cleaned)
        self.save()

    def reset_ms(self, cutoff_ms: int):
        self.load()
        kept = []
        for s, e in self.intervals:
            if e < cutoff_ms:
                kept.append([s, e])
            elif s < cutoff_ms <= e:
                kept.append([s, cutoff_ms - 1])
        self.intervals = self._merge(kept)
        self.save()

    def unprocessed_subintervals(self, beg_ms: int, end_ms: int):
        self.load()
        if beg_ms is None or end_ms is None:
            return []
        if end_ms < beg_ms:
            beg_ms, end_ms = end_ms, beg_ms
        covered = []
        for s, e in self.intervals:
            if e < beg_ms:
                continue
            if s > end_ms:
                break
            covered.append([max(s, beg_ms), min(e, end_ms)])
        covered = self._merge(covered)
        return self._subtract([beg_ms, end_ms], covered)

def filter_new_rows_by_intervals(full_df: pd.DataFrame,
                                 intervals_path: str,
                                 beg_ms: int | None,
                                 end_ms: int | None,
                                 reset_ms: int | None):
    if 'DateTime' not in full_df.columns:
        raise ValueError("❗ Expected 'DateTime' column in RR dataframe.")
    dt_series = pd.to_datetime(full_df['DateTime'], errors='coerce')
    ms_series = (dt_series.astype('int64') // 1_000_000).astype('Int64')
    full_df = full_df.loc[~ms_series.isna()].copy()
    ms_series = ms_series.loc[full_df.index].astype(np.int64)
    if full_df.empty:
        return full_df, []
    min_ms_all, max_ms_all = int(ms_series.min()), int(ms_series.max())
    if beg_ms is None:
        beg_ms = min_ms_all
    if end_ms is None:
        end_ms = max_ms_all

    ledger = IntervalLedger(intervals_path)
    if reset_ms is not None:
        ledger.reset_ms(int(reset_ms))

    gaps = ledger.unprocessed_subintervals(int(beg_ms), int(end_ms))
    if not gaps:
        print("[INFO] ❌ No unprocessed intervals in requested range")
        return full_df.iloc[0:0].copy(), []
    parts, realized = [], []
    for s, e in gaps:
        mask = (ms_series >= s) & (ms_series <= e)
        chunk = full_df.loc[mask]
        if not chunk.empty:
            parts.append(chunk)
            realized.append([int(ms_series.loc[chunk.index].min()), int(ms_series.loc[chunk.index].max())])
    if not parts:
        print("[INFO] 📌 Gaps present but no rows within ranges")
        return full_df.iloc[0:0].copy(), []
    new_df = pd.concat(parts, axis=0).sort_values('DateTime').reset_index(drop=True)
    return new_df, realized

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
''' Helper Functions                                                                  '''
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def create_backup(file_path):
    """Creates a timestamped backup of a file before modification."""
    if not os.path.exists(file_path):
        return
    try:
        dir_name, file_name = os.path.split(file_path)
        name, ext = os.path.splitext(file_name)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(dir_name, f"{name}.{ts}.bak{ext}")
        shutil.copy2(file_path, backup_path)
        print(f"[INFO] 🗄️ Created backup: {os.path.basename(backup_path)}")
    except Exception as e:
        print(f"[WARNING] ❗ Could not create backup for {os.path.basename(file_path)}: {e}")

def create_nan_dict(metric_list):
    return {metric: np.nan for metric in metric_list}

def get_decimals(metric):
    return DECX.get(metric, DEC)

def round_value(x, dec):
    if isinstance(x, dict):
        return {k: round_value(v, dec) for k, v in x.items()}
    elif isinstance(x, tuple):
        return tuple(round_value(i, dec) for i in x)
    elif isinstance(x, (int, float)) and not pd.isnull(x):
        return round(x, dec)
    else:
        return x

def load_data():
    """Loads RR interval data from the path specified in config."""
    rr_file_path = RR_CSV
    print(f"[INFO] 🔄 Loading RR data from {rr_file_path}")
    try:
        df = pd.read_csv(rr_file_path)
        df['DateTime'] = pd.to_datetime(df['DateTime'].astype(float), unit='ms')
        df = df.sort_values('DateTime').reset_index(drop=True)
        return df
    except FileNotFoundError:
        print(f"[ERROR] ⛔ RR data file not found at: {rr_file_path}")
        raise
    except Exception as e:
        print(f"[ERROR] ❌ Failed to load or parse RR data from {rr_file_path}: {e}")
        raise

def compute_time_domain(rr_array, datetime_series=None):
    rr_count = len(rr_array)
    median_RR = np.nanmedian(rr_array)
    min_rr = np.nanmin(rr_array)
    max_rr = np.nanmax(rr_array)
    mxdmn = max_rr - min_rr
    sdr = np.std(rr_array, ddof=1)
    sdsd = np.std(np.diff(rr_array), ddof=1)
    mad = np.median(np.abs(rr_array - median_RR))
    mcv = mad / median_RR * 100 if median_RR != 0 else np.nan
    iqr = np.percentile(rr_array, 75) - np.percentile(rr_array, 25)
    pct20 = np.percentile(rr_array, 20)
    pct80 = np.percentile(rr_array, 80)
    rmssd = np.sqrt(np.mean(np.diff(rr_array)**2))
    pnn50 = (np.sum(np.abs(np.diff(rr_array)) > 50) / (len(rr_array) -1)) * 100
    pnn20 = (np.sum(np.abs(np.diff(rr_array)) > 20) / (len(rr_array) -1)) * 100
    mean_rr = np.nanmean(rr_array)
    cv = (sdr / mean_rr) * 100 if mean_rr != 0 else np.nan
    cvsd = (rmssd / mean_rr) * 100 if mean_rr != 0 and rmssd is not None else np.nan

    dc = ac = np.nan
    if len(rr_array) > 1:
        diffs = np.diff(rr_array)
        decel_anchors = np.where(diffs > 0)[0] + 1
        accel_anchors = np.where(diffs < 0)[0] + 1
        decel_segments = []
        accel_segments = []
        for anchor in decel_anchors:
            if anchor - 5 >= 0 and anchor + 5 < len(rr_array):
                segment = rr_array[anchor - 5:anchor + 6]
                decel_segments.append(segment)
        for anchor in accel_anchors:
            if anchor - 5 >= 0 and anchor + 5 < len(rr_array):
                segment = rr_array[anchor - 5:anchor + 6]
                accel_segments.append(segment)
        if decel_segments:
            decel_avg = np.mean(decel_segments, axis=0)
            dc = (decel_avg[5] + decel_avg[6] - decel_avg[4] - decel_avg[3]) / 4
        if accel_segments:
            accel_avg = np.mean(accel_segments, axis=0)
            ac = -(accel_avg[5] + accel_avg[6] - accel_avg[4] - accel_avg[3]) / 4

    sdann5 = sdnni5 = np.nan
    if datetime_series is not None and len(datetime_series) == len(rr_array):
        try:
            df = pd.DataFrame({'RR': rr_array, 'DateTime': pd.to_datetime(datetime_series)}).set_index('DateTime')
            mean_rr_5min = df['RR'].resample('5T').mean()
            std_rr_5min = df['RR'].resample('5T').std(ddof=1)
            if len(mean_rr_5min.dropna()) > 1:
                sdann5 = mean_rr_5min.std(ddof=1)
                sdnni5 = std_rr_5min.mean()
        except Exception as e:
            print(f"[ALERT] ❌ SDANN/SDNNI failed: {e}")

    return {
        'Count': rr_count,
        'MedRR': median_RR,
        'MinRR': min_rr,
        'MaxRR': max_rr,
        'MxDMn': mxdmn,
        'SDNN': sdr,
        'SDSD': sdsd,
        'MAD': mad,
        'MCV': mcv,
        'IQR': iqr,
        'Q20': pct20,
        'Q80': pct80,
        'RMSSD': rmssd,
        'pNN20': pnn20,
        'pNN50': pnn50,
        'CV%': cv,
        'CVSD': cvsd,
        'DC': dc,
        'AC': ac,
        'SDANN5': sdann5,
        'SDNNI5': sdnni5
    }

def compute_frequency_domain(rr_array, fs=4):
    if rr_array is None or len(rr_array) < 5:
        return create_nan_dict(BM['fd'])
    try:
        rr_seconds = rr_array / 1000.0
        rr_times = np.cumsum(rr_seconds)
        rr_times -= rr_times[0]

        uniform_times = np.arange(0, rr_times[-1], 1/fs)
        if len(uniform_times) < 2:
            return create_nan_dict(BM['fd'])

        interpolated_rr = np.interp(uniform_times, rr_times, rr_array)
        detrended_rr = signal.detrend(interpolated_rr, type='linear')
        nperseg = min(8192, len(detrended_rr))
        if nperseg < fs*2:
            nperseg = len(detrended_rr)
        freqs, psd = signal.welch(detrended_rr, fs=fs, nperseg=nperseg, scaling='density')

        ulf_band, vlf_band = (0, 0.003), (0.003, 0.04)
        lf_band, hf_band = (0.04, 0.15), (0.15, 0.4)

        ulf_power = np.trapz(psd[(freqs>=ulf_band[0])&(freqs<ulf_band[1])],
                             freqs[(freqs>=ulf_band[0])&(freqs<ulf_band[1])])
        vlf_power = np.trapz(psd[(freqs>=vlf_band[0])&(freqs< vlf_band[1])],
                             freqs[(freqs>=vlf_band[0])&(freqs< vlf_band[1])])
        lf_power = np.trapz(psd[(freqs>=lf_band[0])&(freqs< lf_band[1])],
                             freqs[(freqs>=lf_band[0])&(freqs< lf_band[1])])
        hf_power = np.trapz(psd[(freqs>=hf_band[0])&(freqs< hf_band[1])],
                             freqs[(freqs>=hf_band[0])&(freqs< hf_band[1])])

        tp_for_ratios = vlf_power + lf_power + hf_power
        total_power = ulf_power + tp_for_ratios if not np.isnan(ulf_power) else tp_for_ratios
        lf_hf_ratio = lf_power / hf_power if hf_power > 1e-9 else np.nan

        hf_slice = psd[(freqs>=hf_band[0])&(freqs< hf_band[1])]
        hf_freq_slice = freqs[(freqs>=hf_band[0])&(freqs< hf_band[1])]
        hf_peak = hf_freq_slice[np.argmax(hf_slice)] if len(hf_slice)>0 else np.nan

        lf_slice = psd[(freqs>=lf_band[0])&(freqs< lf_band[1])]
        lf_freq_slice = freqs[(freqs>=lf_band[0])&(freqs< lf_band[1])]
        lf_peak = lf_freq_slice[np.argmax(lf_slice)] if len(lf_slice)>0 else np.nan

        vlf_slice = psd[(freqs>=vlf_band[0])&(freqs< vlf_band[1])]
        vlf_freq_slice = freqs[(freqs>=vlf_band[0])&(freqs< vlf_band[1])]
        vlf_peak = vlf_freq_slice[np.argmax(vlf_slice)] if len(vlf_slice)>0 else np.nan

        lf_rel = lf_power/total_power*100 if total_power>1e-9 else np.nan
        hf_rel = hf_power/total_power*100 if total_power>1e-9 else np.nan
        vlf_rel = vlf_power/total_power*100 if total_power>1e-9 else np.nan

        return {
            'Total Power': total_power,
            'ULF': ulf_power,
            'VLF': vlf_power,
            'VLF%': vlf_rel,
            'VLFp': vlf_peak,
            'LFHF': lf_hf_ratio,
            'LF': lf_power,
            'LF%': lf_rel,
            'LFp': lf_peak,
            'HF': hf_power,
            'HF%': hf_rel,
            'HFp': hf_peak,
        }
    except Exception as e:
        print(f"[ERROR] ⛔ Error in compute_frequency_domain: {e}")
        return create_nan_dict(BM['fd'])

def compute_non_linear_metrics(rr_array):
    try:
        pe = ant.perm_entropy(rr_array, normalize=True)
    except Exception as e:
        print(f"[WARNING] ❌ PE failed: {e}")
        pe = np.nan
    try:
        se = ant.sample_entropy(rr_array, order=2, metric='chebyshev')
    except Exception as e:
        print(f"[WARNING] ❌ SE failed: {e}")
        se = np.nan
    try:
        split = len(rr_array)//2
        df1 = ant.detrended_fluctuation(rr_array[:split]) if split>10 else np.nan
        df2 = ant.detrended_fluctuation(rr_array[split:]) if (len(rr_array)-split)>10 else np.nan
    except Exception as e:
        print(f"[WARNING] ❌ DFA failed: {e}")
        df1 = df2 = np.nan
    return {'PermEn': pe, 'SampEn': se, 'MFE': np.nan, 'CorrDim': np.nan, 'DFA 𝛼1': df1, 'DFA 𝛼2': df2}

def compute_geometric_metrics(rr_array):
    if rr_array is None or len(rr_array)<2:
        return create_nan_dict(BM['gm'])
    metrics = create_nan_dict(BM['gm'])
    try:
        diff_rr = np.diff(rr_array)
        metrics['SD1'] = np.sqrt(0.5*np.var(diff_rr, ddof=1))
        metrics['SD2'] = np.sqrt(2*np.var(rr_array, ddof=1)-0.5*np.var(diff_rr, ddof=1))
        metrics['SD1SD2'] = metrics['SD1']/metrics['SD2'] if metrics['SD2']>1e-9 else np.nan
        metrics['CSI'] = metrics['SD2']/metrics['SD1'] if metrics['SD1']>1e-9 else np.nan
        metrics['CVI'] = np.log10(metrics['SD1']*metrics['SD2']) if metrics['SD1']>0 and metrics['SD2']>0 else np.nan
    except Exception as e:
        print(f"[WARNING] ❌ Poincare metrics failed: {e}")
        metrics.update({'SD1':np.nan,'SD2':np.nan,'SD1SD2':np.nan,'CSI':np.nan,'CVI':np.nan})
    try:
        hist, edges = np.histogram(rr_array, bins='auto')
        if len(hist)>0 and hist.max()>0:
            metrics['HTI'] = len(rr_array)/hist.max()
            half = hist.max()/2
            idx = np.where(hist>=half)[0]
            if len(idx)>0:
                start, end = edges[idx[0]], edges[idx[-1]+1]
                metrics['TINN'] = end-start
            else:
                metrics['TINN']=np.nan
        else:
            metrics['HTI']=metrics['TINN']=np.nan
        hist50,_=np.histogram(rr_array, bins=np.arange(rr_array.min(),rr_array.max()+50,50))
        metrics['AMo50%'] = hist50.max()/len(rr_array)*100 if len(hist50)>0 and hist50.max()>0 else np.nan
    except Exception as e:
        print(f"[WARNING] ❌ Histogram metrics failed: {e}")
        metrics['HTI']=metrics['TINN']=metrics['AMo50%']=np.nan
    return metrics

def get_minimum_intervals(timeframe):
    return MIN_INT.get(timeframe, 250)

def compute_hrv_metrics(rr_window, start_datetime, end_datetime, timeframe):
    min_beats = get_minimum_intervals(timeframe)

    if rr_window is None or len(rr_window) < min_beats:
        print(f"[WARNING] ⚠️ Window {timeframe} from {start_datetime} to {end_datetime} contains only {len(rr_window) if rr_window is not None else 0} intervals. Minimum required: {min_beats}. Skipping.")

        all_keys = set().union(*BM.values())
        nan_dict = create_nan_dict(list(all_keys))
        filtered = {k:v for k,v in nan_dict.items() if k in TF.get(timeframe, set())}

        return {'StartTime': start_datetime, 'EndTime': end_datetime, **filtered}

    rr_array = np.array(rr_window)
    dt_range=None

    try:
        dt_range = pd.date_range(start=pd.to_datetime(start_datetime), end=pd.to_datetime(end_datetime), periods=len(rr_array))
    except Exception:
        dt_range=None

    td = compute_time_domain(rr_array, datetime_series=dt_range)
    fd = compute_frequency_domain(rr_array)
    nl = compute_non_linear_metrics(rr_array)
    gm = compute_geometric_metrics(rr_array)

    metrics = {**td, **fd, **nl, **gm}
    filtered = {k:v for k,v in metrics.items() if k in TF.get(timeframe, set())}

    return {'StartTime': start_datetime, 'EndTime': end_datetime, **filtered}

def process_window(args):
    window_data, start_dt, end_dt, tf = args
    print(f"[INFO] 🔄 Processing {tf} window ({len(window_data)} intervals) from {start_dt} to {end_dt}")
    try:
        return compute_hrv_metrics(window_data, start_dt, end_dt, tf)
    except Exception as e:
        print(f"[ERROR] ⛔ Error processing window {tf} ({start_dt}-{end_dt}): {e}")
        all_keys = set().union(*BM.values())
        nan_dict = create_nan_dict(list(all_keys))
        filtered = {k:v for k,v in nan_dict.items() if k in TF.get(tf, set())}
        return {'StartTime': start_dt, 'EndTime': end_dt, **filtered}

def calculate_hrv_metrics_dynamic_windows(rr_series, datetime_series, window_duration=5,
                                          timeframes=TIMEFRAMES, num_processes=mp.cpu_count()):
    try:
        datetime_series = pd.to_datetime(datetime_series).tolist()
    except Exception as e:
        print(f"[ERROR] ❌ Failed to convert datetime_series to datetime objects: {e}")
        return {tf:[] for tf in timeframes}
    if not datetime_series:
        print("[ERROR] ⛔ datetime_series is empty. Cannot process windows.")
        return {tf:[] for tf in timeframes}

    hrv_raw = {tf:[] for tf in timeframes}
    print(f"[INFO] 🔄 Initiating parallel computation with {num_processes} processes...")

    deltas = {
        '5M': timedelta(minutes=window_duration), '15M': timedelta(minutes=15), '30M': timedelta(minutes=30), '1H': timedelta(hours=1),
        '2H': timedelta(hours=2), '3H': timedelta(hours=3), '6H': timedelta(hours=6), '12H': timedelta(hours=12), '24H': timedelta(hours=24)
    }

    with mp.Pool(processes=num_processes) as pool:
        args_list=[]
        for tf in timeframes:
            print(f"\n[INFO] 🔄 Preparing windows for timeframe: {tf}")
            delta = deltas.get(tf)
            if delta is None:
                print(f"[WARNING] ❗ Unknown timeframe '{tf}'. Skipping.")
                continue
            current_start = datetime_series[0]
            current_end = current_start + delta
            buffer = []
            start_dt = current_start
            for rr, dt in zip(rr_series, datetime_series):
                if dt < current_start:
                    continue
                if dt < current_end:
                    buffer.append(rr)
                else:
                    if len(buffer)>=get_minimum_intervals(tf):
                        args_list.append((buffer.copy(), start_dt, current_end, tf))
                    else:
                        if buffer:
                            print(f"[DEBUG] ❗ Skipping {tf} window ({start_dt} to {current_end}): insufficient data ({len(buffer)}<{get_minimum_intervals(tf)})")
                    while dt>=current_end:
                        current_start+=delta
                        current_end+=delta
                    start_dt=current_start
                    buffer=[rr]
        if buffer and len(buffer) >= get_minimum_intervals(tf):
            args_list.append((buffer.copy(), start_dt, current_end, tf))
        elif buffer:
            print(f"[DEBUG] ❗ Skipping final {tf} window ({start_dt} to {current_end}): insufficient data ({len(buffer)}<{get_minimum_intervals(tf)})")

        for tf in timeframes:
            print(f"[INFO] 📌 Prepared {len([a for a in args_list if a[3]==tf])} windows for {tf}.")

        if args_list:
            print(f"\n[INFO] 🔄 Processing {len(args_list)} windows across all timeframes...")
            try:
                results = pool.map(process_window, args_list)
                for res in results:
                    if res and 'StartTime' in res:
                        dur = res['EndTime'] - res['StartTime']
                        matched = None
                        for t,d in deltas.items():
                            if abs(dur.total_seconds()-d.total_seconds())<1:
                                matched=t;break
                        if matched:
                            hrv_raw[matched].append(res)
                        else:
                            print(f"[WARNING] ❗ Could not match result window duration {dur} to a timeframe. Discarding.")
            except Exception as e:
                print(f"[ERROR] ❌ Parallel window processing failed: {e}")
        else:
            print("[INFO] 📌 No valid windows found to process")

    combined = {tf: hrv_raw[tf] for tf in timeframes}
    print("[INFO] ✅ Dynamic window processing completed")
    return combined

def compute_summary_statistics(hrv_metrics_df):
    if hrv_metrics_df is None or hrv_metrics_df.empty:
        return pd.DataFrame()
    numeric_df = hrv_metrics_df.select_dtypes(include=[np.number])
    if numeric_df.empty:
        return pd.DataFrame()
    summary = numeric_df.describe().transpose()
    variance = numeric_df.var()
    median = numeric_df.median()
    mode1 = numeric_df.mode().iloc[0] if not numeric_df.mode().empty else pd.Series(dtype='float64')
    summary['Variance'] = variance
    summary['Median'] = median
    summary['Mode'] = mode1
    summary = summary.applymap(lambda x: round(x,1) if not pd.isnull(x) else x)
    summary = summary.applymap(lambda x: int(x) if isinstance(x,float) and x.is_integer() else x)
    return summary

# ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼
# ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼
#
# THIS FUNCTION HAS BEEN REPLACED TO CORRECTLY HANDLE DIRECT FILE PATHS
#
def find_excel_workbook(exact_file_path):
    """
    Returns the provided file path. If the file does not exist,
    it creates a new, empty workbook at that path.
    """
    if not os.path.exists(exact_file_path):
        # Create the directory if it doesn't exist
        dir_name = os.path.dirname(exact_file_path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        # Create an empty workbook
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(exact_file_path)
        print(f"[INFO] 📌 Workbook not found. Created new one: {os.path.basename(exact_file_path)}")
    else:
        print(f"[INFO] 📌 Found existing workbook: {os.path.basename(exact_file_path)}")

    return exact_file_path
#
# ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲
# ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲

def update_excel_workbook(hrv_metrics_dfs, existing_file_path, workbook_mode='metrics'):
    """
    Writes timeframe sheets and a single stacked Summary sheet.
    This version writes to a temporary file first and then renames it to prevent
    I/O timeouts and validation errors.
    """
    # NOTE: EXCLUDE set and associated logic has been removed as requested.

    existing_data = {}
    try:
        with pd.ExcelFile(existing_file_path, engine="openpyxl") as xls:
            sheets = set(xls.sheet_names)
            for tf in TIMEFRAMES:
                sn = tf if tf in sheets else (f"{tf} Metrics" if f"{tf} Metrics" in sheets else None)
                if sn:
                    try:
                        df = pd.read_excel(xls, sheet_name=sn, dtype={'StartTime':object,'EndTime':object})
                        if 'StartTime' in df.columns:
                            df['StartTime'] = pd.to_datetime(df['StartTime'], errors='coerce')
                            df = df.dropna(subset=['StartTime'])
                        if 'EndTime' in df.columns:
                            df['EndTime'] = pd.to_datetime(df['EndTime'], errors='coerce')
                        existing_data[tf] = df
                        print(f"[INFO] 📌 Loaded existing sheet '{sn}' ({len(df)} rows)")
                    except Exception as e:
                        print(f"[WARNING] ❗ Could not read sheet '{sn}': {e}")
    except Exception as e:
        print(f"[WARNING] ❗ Could not open existing workbook with openpyxl: {e}")
        existing_data = {}

    combined = {}
    for tf in TIMEFRAMES:
        new_df = hrv_metrics_dfs.get(tf)
        old_df = existing_data.get(tf)
        use_df = pd.DataFrame()

        if new_df is not None and not new_df.empty:
            use_df = new_df.copy()
            if 'StartTime' in use_df.columns: use_df['StartTime'] = pd.to_datetime(use_df['StartTime'], errors='coerce')
            if 'EndTime' in use_df.columns: use_df['EndTime'] = pd.to_datetime(use_df['EndTime'], errors='coerce')
            use_df = use_df.dropna(subset=['StartTime']) if 'StartTime' in use_df.columns else use_df

        if old_df is not None and not old_df.empty and not use_df.empty:
            both = pd.concat([old_df, use_df], ignore_index=True)
            if 'StartTime' in both.columns:
                both = both.drop_duplicates(subset=['StartTime'], keep='last').sort_values('StartTime')
            combined[tf] = both
        elif use_df is not None and not use_df.empty:
            combined[tf] = use_df.sort_values('StartTime') if 'StartTime' in use_df.columns else use_df
        elif old_df is not None and not old_df.empty:
            combined[tf] = old_df
        else:
            combined[tf] = pd.DataFrame()

    def select_columns(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty: return df
        cols = df.columns.tolist()
        front = [c for c in ['StartTime','EndTime'] if c in cols]
        numeric = [c for c in cols if c not in front]
        if workbook_mode == 'stats':
            keep_numeric = [c for c in numeric if c in STATS]
        else:
            keep_numeric = [c for c in numeric if c not in STATS]
        return df[front + keep_numeric] if keep_numeric else pd.DataFrame(columns=front)

    dir_name, file_name = os.path.split(existing_file_path)
    temp_file_path = os.path.join(dir_name, f"~temp_{os.getpid()}_{file_name}")

    try:
        wrote_any = False
        with pd.ExcelWriter(temp_file_path, engine='openpyxl', mode='w') as writer:
            for tf in TIMEFRAMES:
                df_tf = select_columns(combined.get(tf))
                if df_tf is None or df_tf.empty: continue

                if 'StartTime' in df_tf.columns and pd.api.types.is_datetime64_any_dtype(df_tf['StartTime']):
                    df_tf['StartTime'] = df_tf['StartTime'].apply(lambda dt: dt.strftime(DT_XL_FORMAT) if pd.notna(dt) else '')
                if 'EndTime' in df_tf.columns and pd.api.types.is_datetime64_any_dtype(df_tf['EndTime']):
                    df_tf['EndTime'] = df_tf['EndTime'].apply(lambda dt: dt.strftime(DT_XL_FORMAT) if pd.notna(dt) else '')

                for col in df_tf.select_dtypes(include=np.number).columns:
                    df_tf[col] = df_tf[col].apply(lambda v: round(float(v), get_decimals(col)) if pd.notna(v) else v)

                df_tf.to_excel(writer, sheet_name=tf, index=False)
                wrote_any = True

            if not wrote_any:
                pd.DataFrame({"Info":["No data for this mode"]}).to_excel(writer, sheet_name="Summary", index=False)

        wb = load_workbook(temp_file_path)
        if 'Summary' in wb.sheetnames: del wb['Summary']
        ws = wb.create_sheet('Summary')

        cur_row, title_font, center, dbl = 1, Font(name='Aptos Serif', bold=True, size=20), Alignment(horizontal='center', vertical='center'), Border(bottom=Side(style='double'))

        for tf in TIMEFRAMES:
            base = select_columns(combined.get(tf))
            if base is None or base.empty: continue
            sdf = base.drop(columns=['StartTime','EndTime'], errors='ignore')
            if sdf.empty: continue

            # Clean and coerce to numeric to avoid errors from stray whitespace/strings
            sdf = sdf.copy()
            for c in sdf.columns:
                if pd.api.types.is_object_dtype(sdf[c]):
                    # Strip whitespace and turn empty strings into NaN
                    sdf[c] = sdf[c].astype(str).str.strip().replace({'': np.nan})
                if not pd.api.types.is_numeric_dtype(sdf[c]):
                    sdf[c] = pd.to_numeric(sdf[c], errors='coerce')
            # Keep only numeric columns for summary statistics
            sdf = sdf.select_dtypes(include=[np.number])
            if sdf.empty: continue

            desc = sdf.describe().transpose()
            desc['Variance'] = sdf.var(numeric_only=True)
            desc['Median'] = sdf.median(numeric_only=True)
            mode_df = sdf.mode()
            mode_row = mode_df.iloc[0] if not mode_df.empty else pd.Series(dtype=float)
            desc['Mode'] = mode_row
            desc = desc.reset_index().rename(columns={'index':'Metric'})

            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=12)
            tcell = ws.cell(row=cur_row, column=1, value=tf)
            tcell.font, tcell.alignment = title_font, center
            for c in range(1, 13): ws.cell(row=cur_row, column=c).border = dbl

            r = cur_row + 1
            for j, col in enumerate(desc.columns, start=1):
                ws.cell(row=r, column=j, value=col)
            for i in range(len(desc)):
                for j, col in enumerate(desc.columns, start=1):
                    ws.cell(row=r+1+i, column=j, value=desc.iloc[i, j-1])
            cur_row = r + 1 + len(desc) + 3

        wb.save(temp_file_path)
        shutil.move(temp_file_path, existing_file_path)

    except Exception as e:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        raise e

    print(f"[INFO] ✅ Successfully updated Excel workbook: {os.path.basename(existing_file_path)}")
    return combined

def rename_excel_workbook(excel_path, combined_dfs, base_label=None):
    try:
        min_date=max_date=None
        for tf, df in combined_dfs.items():
            if df is None or df.empty:
                continue
            st=pd.to_datetime(df['StartTime'], errors='coerce') if 'StartTime' in df.columns else None
            et=pd.to_datetime(df['EndTime'], errors='coerce') if 'EndTime' in df.columns else None
            if st is not None and not st.dropna().empty:
                dmin=st.min()
                if min_date is None or dmin<min_date:
                    min_date=dmin
            curr_max=None
            if et is not None and not et.dropna().empty:
                curr_max=et.max()
            elif st is not None and not st.dropna().empty:
                curr_max=st.max()
            if curr_max is not None and (max_date is None or curr_max>max_date):
                max_date=curr_max

        if min_date is None or max_date is None:
            print("[WARNING] ⛔ Could not determine date range for renaming")
            return excel_path

        s_str=min_date.strftime(DT_XL_FILENAME)
        e_str=max_date.strftime(DT_XL_FILENAME)
        dirn=os.path.dirname(excel_path)
        base = base_label or XL_BASE
        new_fn = f"{base} {s_str} - {e_str}.xlsx"
        new_path=os.path.join(dirn, new_fn)
        if os.path.abspath(excel_path)!=os.path.abspath(new_path):
            if os.path.exists(new_path):
                print(f"[INFO] 🔄 Removing existing file with target name: {new_fn}")
                os.remove(new_path)
            os.rename(excel_path, new_path)
            print(f"[INFO] ✅ Excel workbook renamed to: {new_fn}")
            return new_path
        return excel_path
    except Exception as e:
        print(f"[ERROR] ⛔ Error during Excel workbook renaming: {e}")
        return excel_path

def entropic_data(rr_data_path):
    print("[INFO] ————————————————————— Entropic —————————————————————")
    print(f"[INFO] 🔄 Loading RR data from {rr_data_path}...")
    df=pd.read_csv(rr_data_path)
    df['DateTime']=pd.to_datetime(df['DateTime'].astype(float), unit='ms')
    df=df.sort_values('DateTime').reset_index(drop=True)
    print(f"[INFO] 🔄 Loading & processing {len(df)} RR intervals...")
    return df

def compute_entropic_metrics(rr_array, dimension):
    metrics={'MFE':np.nan,'CorrDim':np.nan}
    if rr_array is None or len(rr_array)<dimension*5:
        return metrics
    try:
        mfe_result=nk.entropy_multiscale(rr_array, dimension=dimension, scale=3, fuzzy=True)
        if isinstance(mfe_result, tuple) and mfe_result:
            metrics['MFE']=round(float(mfe_result[0]),4)
        elif isinstance(mfe_result,(int,float)):
            metrics['MFE']=round(float(mfe_result),4)
    except Exception as e:
        print(f"[WARNING] ❌ MFE failed: {e}")
    try:
        cd=nolds.corr_dim(rr_array, emb_dim=dimension)
        metrics['CorrDim']=round(float(cd),3) if abs(cd)>1e-10 else np.nan
    except Exception as e:
        print(f"[WARNING] ❌ CorrDim failed: {e}")
    return metrics

def segment_windows(rr_series, datetime_series, delta, min_interval_count=0):
    windows=[]
    if not datetime_series:
        print("[WARNING] ⚠️ Empty datetime series for segment_windows")
        return windows
    try:
        datetime_series=list(pd.to_datetime(datetime_series))
    except Exception as e:
        print(f"[ERROR] ❌ segment_windows datetime conversion failed: {e}")
        return windows

    curr_start=datetime_series[0]
    curr_end=curr_start+delta
    buf=[]
    for rr, dt in zip(rr_series, datetime_series):
        if dt<curr_start:
            continue
        if dt<curr_end:
            buf.append(rr)
        else:
            if len(buf)>=min_interval_count:
                windows.append((curr_start, curr_end, buf.copy()))
            while dt>=curr_end:
                curr_start+=delta
                curr_end+=delta
            buf=[rr]
    if buf and len(buf)>=min_interval_count:
        windows.append((curr_start, curr_end, buf.copy()))
    return windows

def save_checkpoint(df, interval, checkpoint_dir):
    if df is None or df.empty:
        print(f"[INFO] No data to save entropic checkpoint for {interval}.")
        return
    path=os.path.join(checkpoint_dir, f"checkpoint_{interval}.parquet")
    try:
        os.makedirs(checkpoint_dir, exist_ok=True)
        df.to_parquet(path, engine='pyarrow', index=False)
        print(f"[INFO] 💾 Entropic checkpoint saved for {interval}: {path} ({len(df)} rows)")
    except Exception as e:
        print(f"[ERROR] ❌ Failed to save checkpoint for {interval}: {e}")

def load_checkpoints(checkpoint_dir, intervals):
    loaded, last_proc = {}, {}
    os.makedirs(checkpoint_dir, exist_ok=True)
    for interval in intervals:
        file=os.path.join(checkpoint_dir, f"checkpoint_{interval}.parquet")
        if os.path.exists(file):
            try:
                df=pd.read_parquet(file, engine='pyarrow')
                if 'StartTime' in df.columns:
                    df['StartTime']=pd.to_datetime(df['StartTime'], errors='coerce')
                    df=df.dropna(subset=['StartTime'])
                loaded[interval]=df
                if not df.empty:
                    last_proc[interval]=df['StartTime'].max()
                    print(f"[INFO] 🔄 Loaded checkpoint for {interval}. Resuming after {last_proc[interval]} ({len(df)} rows)")
                else:
                    print(f"[INFO] 🔄 Loaded empty checkpoint for {interval}.")
            except Exception as e:
                print(f"[WARNING] ❌ Failed to load checkpoint {file}: {e}")
                loaded[interval]=pd.DataFrame(columns=['StartTime','EndTime','MFE','CorrDim'])
        else:
            print(f"[INFO] 📌 No checkpoint found for {interval}.")
            loaded[interval]=pd.DataFrame(columns=['StartTime','EndTime','MFE','CorrDim'])
    return loaded, last_proc

def clear_checkpoints(checkpoint_dir, intervals):
    print("[INFO] 🔄 Clearing checkpoint files...")
    cnt=0
    for interval in intervals:
        file=os.path.join(checkpoint_dir, f"checkpoint_{interval}.parquet")
        if os.path.exists(file):
            try:
                os.remove(file)
                cnt+=1
            except Exception as e:
                print(f"[WARNING] ❌ Failed to remove checkpoint file {file}: {e}")
    print(f"[INFO] 🧹 Cleared {cnt} checkpoint files from {checkpoint_dir}")

def format_datetime_for_excel(dt):
    if isinstance(dt, datetime):
        try:
            return dt.strftime(DT_XL_FORMAT)
        except:
            return ''
    return dt

def find_latest_excel(directory, base_label=None):
    base = base_label
    pattern=os.path.join(directory, f"{base}*.xlsx")
    files=glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"⚠️ No Excel files found matching '{base}*.xlsx' in {directory}")
    latest=max(files, key=os.path.getmtime)
    print(f"[INFO] ✅ Found latest Excel file ⟨{os.path.basename(latest)}⟩")
    return latest

def update_excel_with_metrics(excel_file, entropic_results, intervals):
    """
    [ROBUST VERSION]: Merges entropic metrics by reading the entire workbook into
    memory, performing the merge, and writing to a new temporary file before
    replacing the original. This avoids all in-place modification bugs.
    """
    print(f"[INFO] 🔄 Merging entropic metrics into {os.path.basename(excel_file)}...")

    if not os.path.exists(excel_file):
        print(f"[ERROR] ❌ Workbook not found at {excel_file}. Cannot merge.")
        return False

    # --- STAGE 1: Read the ENTIRE existing workbook into memory ---
    try:
        # Read all sheets into a dictionary of DataFrames. This preserves untouched sheets.
        all_sheets_data = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        print(f"[INFO] 📌 Loaded {len(all_sheets_data)} sheets from the existing workbook.")
    except Exception as e:
        print(f"[ERROR] ❌ Failed to read the existing Excel file '{os.path.basename(excel_file)}': {e}")
        return False

    had_errors = False

    # --- STAGE 2: Perform the merge on the relevant sheets IN MEMORY ---
    for interval in intervals:
        if interval not in entropic_results or entropic_results[interval].empty:
            continue # No new entropic data for this interval

        if interval not in all_sheets_data:
            print(f"[WARNING] ❗ Sheet '{interval}' not found in workbook. Skipping merge for this interval.")
            continue

        try:
            df_existing = all_sheets_data[interval]
            df_new = entropic_results[interval]

            if df_existing.empty or 'StartTime' not in df_existing.columns:
                print(f"[WARNING] ❗ Sheet '{interval}' is empty or missing 'StartTime'. Skipping.")
                continue

            # Perform the merge using the robust string-key method
            df_existing = df_existing.copy()
            df_new = df_new.copy()

            df_new['StartTime'] = pd.to_datetime(df_new['StartTime'], errors='coerce')
            df_new.dropna(subset=['StartTime'], inplace=True)
            df_new['StartTime_Key'] = df_new['StartTime'].apply(lambda dt: dt.strftime(DT_XL_FORMAT) if pd.notna(dt) else None)

            for col in ['MFE', 'CorrDim']:
                if col in df_existing.columns:
                    df_existing.drop(columns=[col], inplace=True)

            cols_to_use = [col for col in ['StartTime_Key', 'MFE', 'CorrDim'] if col in df_new.columns]

            df_merged = pd.merge(
                df_existing, df_new[cols_to_use],
                left_on='StartTime', right_on='StartTime_Key', how='left'
            )

            if 'StartTime_Key' in df_merged.columns:
                df_merged.drop(columns=['StartTime_Key'], inplace=True)

            # Reorder columns to ensure new ones are at the end
            existing_cols_order = [c for c in df_existing.columns if c not in ['MFE', 'CorrDim']]
            new_cols_order = [c for c in ['MFE', 'CorrDim'] if c in df_merged.columns]
            final_cols = existing_cols_order + new_cols_order
            df_merged = df_merged[final_cols]

            # IMPORTANT: Update the dictionary with the modified DataFrame
            all_sheets_data[interval] = df_merged
            print(f"[INFO] ✅ Merged entropic data for sheet '{interval}'.")

        except Exception as e:
            print(f"[ERROR] ❌ An unexpected error occurred while processing sheet '{interval}': {e}")
            traceback.print_exc()
            had_errors = True
            break

    if had_errors:
        print("[ERROR] ❌ Aborting Excel update due to processing errors.")
        return False

    # --- STAGE 3: Write the ENTIRE workbook to a temporary file ---
    dir_name, file_name = os.path.split(excel_file)
    temp_file_path = os.path.join(dir_name, f"~temp_entropic_{os.getpid()}_{file_name}")

    try:
        with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
            for sheet_name, df_to_write in all_sheets_data.items():
                # Apply rounding to numeric columns before writing
                for col in df_to_write.select_dtypes(include=np.number).columns:
                    df_to_write[col] = df_to_write[col].apply(lambda v: round(float(v), get_decimals(col)) if pd.notna(v) else v)
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)

        # --- STAGE 4: Replace the original file with the new one ---
        shutil.move(temp_file_path, excel_file)

        print(f"[INFO] ✅ Successfully rebuilt and saved the workbook: {os.path.basename(excel_file)}")
        return True # Return SUCCESS

    except Exception as e:
        print(f"[ERROR] ❌ Failed to write and replace the final workbook: {e}")
        traceback.print_exc()
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path) # Clean up failed temp file
        return False # Return FAILURE

def run_entropic_metrics():
    rr_data_path=RR_CSV
    output_dir=OUTPUTS_DIR
    checkpoint_dir=CHECKPOINT_DIR
    entropic_intervals=ENTROPIC_INTERVALS

    print("[INFO] ✨ Starting Entropic Calculations...")
    entropic_results, last_proc = load_checkpoints(checkpoint_dir, entropic_intervals.keys())
    try:
        df=entropic_data(rr_data_path)
    except Exception as e:
        print(f"[ERROR] ❌ Failed to load RR data: {e}")
        return

    rr_series=df['RR'].tolist()
    datetime_series=df['DateTime'].tolist()
    all_list = {}
    for interval in entropic_intervals:
        existing = entropic_results.get(interval)
        if isinstance(existing, pd.DataFrame):
            current_records = existing.to_dict('records')
        elif isinstance(existing, list):
            current_records = existing
        elif existing is None:
            current_records = []
        else:
            try:
                current_records = list(existing)
            except Exception:
                current_records = []
        all_list[interval] = current_records

    for interval, params in entropic_intervals.items():
        print(f"\n[INFO] ⏳ Entropic: Processing interval: {interval}")
        delta=params['delta']; dimension=params['dimension']; min_int=params['min_intervals']; cp=params['cp_fractions']
        windows=segment_windows(rr_series, datetime_series, delta, min_interval_count=min_int)
        print(f"[INFO] Entropic: Created {len(windows)} valid windows for {interval}.")
        resume=last_proc.get(interval)
        new_count=0
        num_w=len(windows)
        checkpoints=set(min(max(0,int(num_w*frac)-1), num_w-1) for frac in sorted(set(cp)))
        current=all_list[interval]
        for idx,(st,et,wr) in enumerate(windows):
            if resume is not None and st<=resume:
                continue
            metrics={'MFE':np.nan,'CorrDim':np.nan}
            if len(wr)<10:
                print(f"[ALERT] 🔴 Window {idx+1} for {interval} has only {len(wr)} points (<10) and will be skipped")
            else:
                metrics=compute_entropic_metrics(np.array(wr), dimension)
            current.append({'StartTime':st, 'EndTime':et, **metrics})
            new_count+=1
            if idx in checkpoints:
                save_checkpoint(pd.DataFrame(current), interval, checkpoint_dir)
        final_df=pd.DataFrame(current)
        save_checkpoint(final_df, interval, checkpoint_dir)
        entropic_results[interval]=final_df
        print(f"[INFO] ✅ Entropic: {interval} processed {new_count} new windows. Total: {len(final_df)}")

    try:
        # Use the configured absolute path directly; create if missing
        excel_file = find_excel_workbook(XL_MET)
        create_backup(excel_file)
    except Exception as e:
        print(f"[ERROR] ⛔ Could not prepare Excel file: {e}")
        print(f"[ERROR] ❌ Entropic calculations failed")
        return

    status=update_excel_with_metrics(excel_file, entropic_results, entropic_intervals.keys())
    if status:
        print("[INFO] ✅ Entropic calculations completed")
        print("[INFO] ✅ Excel update completed")
        clear_checkpoints(checkpoint_dir, entropic_intervals.keys())
    else:
        print("[WARNING] 🔴 Entropic calculations completed but the Excel update FAILED.")
        print("[WARNING] 💾 Checkpoints have been retained. Please resolve the Excel issue and re-run.")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
''' MAIN ENTRY POINT — UPDATED WITH CHECKPOINTING                                     '''
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def main():
    parser = argparse.ArgumentParser(description="HRV Analysis Script")
    parser.add_argument("--h", type=int, default=None, help="Limit processing to the first H rows of new data.")
    parser.add_argument("--s", type=int, nargs="*", default=None,
                        help=f"Sequent script numbers to execute upon completion.")
    parser.add_argument("--a", action="store_true", help="Run Entropic (MFE, CorrDim) calculations in addition to main HRV analysis.")
    parser.add_argument("--e", action="store_true", help="Run ONLY Entropic calculations, skipping main HRV processing.")
    parser.add_argument("--beg-ms", type=int, default=None, help="Only process data with timestamps >= this ms since epoch.")
    parser.add_argument("--end-ms", type=int, default=None, help="Only process data with timestamps <= this ms since epoch.")
    parser.add_argument("--reset-ms", type=int, default=None, help="Discard coverage >= this ms in the interval ledger before processing.")
    parser.add_argument("--fresh", action="store_true", help="Clear existing checkpoints and force a full recalculation.")
    args = parser.parse_args()

    if args.e:
        print("[INFO] === Running in Entropic Only Mode ===")
        run_entropic_metrics()
        print("[INFO] === Entropic Only Mode Finished ===")
        return

    print("[INFO] === Initiating Standard HRV Analysis ===")
    start_time = datetime.now()

    checkpoint_dir = CHECKPOINT_DIR
    os.makedirs(checkpoint_dir, exist_ok=True)
    gaps_checkpoint_path = os.path.join(checkpoint_dir, "hrv_checkpoint_gaps.json")

    if args.fresh:
        print("[INFO] 🧹 --fresh flag used. Clearing existing checkpoints...")
        for tf in TIMEFRAMES:
            cp_path = os.path.join(checkpoint_dir, f"hrv_checkpoint_{tf}.parquet")
            if os.path.exists(cp_path): os.remove(cp_path)
        if os.path.exists(gaps_checkpoint_path): os.remove(gaps_checkpoint_path)

    hrv_metrics_dfs = {}
    gaps = []
    checkpoint_found = False

    try:
        configured_tfs = [tf for tf in TIMEFRAMES if tf in TF]
        all_cps_exist = all(os.path.exists(os.path.join(checkpoint_dir, f"hrv_checkpoint_{tf}.parquet")) for tf in configured_tfs)

        if all_cps_exist and os.path.exists(gaps_checkpoint_path):
            print("[INFO] 💾 Found existing checkpoints. Loading data...")
            for tf in configured_tfs:
                cp_path = os.path.join(checkpoint_dir, f"hrv_checkpoint_{tf}.parquet")
                hrv_metrics_dfs[tf] = pd.read_parquet(cp_path)
            with open(gaps_checkpoint_path, 'r') as f:
                gaps = json.load(f)
            print("[INFO] ✅ Checkpoint data loaded. Skipping calculations.")
            checkpoint_found = True
    except Exception as e:
        print(f"[WARNING] ❗ Failed to load checkpoints: {e}. Proceeding with full calculation.")
        hrv_metrics_dfs = {}

    if not checkpoint_found:
        try:
            full_df = load_data()
            print(f"[INFO] 📌 Loaded RR data with {len(full_df)} rows")
        except Exception as e:
            print(f"[ERROR] ❌ Failed to load RR data: {e}")
            return

        try:
            new_df, gaps = filter_new_rows_by_intervals(full_df, str(INTERVALS_JSON), args.beg_ms, args.end_ms, args.reset_ms)
            print(f"[INFO] 📌 Filtered data: {len(new_df)} unprocessed rows in {len(gaps)} interval(s).")

            if args.h is not None and args.h < len(new_df):
                print(f"[INFO] 📌 Limiting to first {args.h} rows.")
                new_df = new_df.head(args.h)

            if new_df.empty:
                print("[INFO] ⛔ No unprocessed data to analyze.")
                if args.a:
                    print("\n[INFO] === Proceeding to Entropic Calculation ===")
                    run_entropic_metrics()
                else:
                    print("[INFO] 🛑 Exiting.")
                return
        except Exception as e:
            print(f"[ERROR] ❌ Interval filtering failed: {e}")
            return

        print(f"[INFO] 🔄 Processing {len(new_df)} new rows for standard HRV...")
        hrv_metrics = calculate_hrv_metrics_dynamic_windows(
            rr_series=new_df['RR'].tolist(),
            datetime_series=new_df['DateTime'].tolist(),
            window_duration=5,
            timeframes=TIMEFRAMES,
            num_processes=mp.cpu_count()
        )
        hrv_metrics_dfs = {tf: pd.DataFrame(metrics) for tf, metrics in hrv_metrics.items() if metrics}

        print("[INFO] 💾 Saving calculation checkpoints...")
        for tf, df in hrv_metrics_dfs.items():
            cp_path = os.path.join(checkpoint_dir, f"hrv_checkpoint_{tf}.parquet")
            df.to_parquet(cp_path, index=False)
        with open(gaps_checkpoint_path, 'w') as f:
            json.dump(gaps, f)

    try:
        excel_path_met = find_excel_workbook(XL_MET)
        create_backup(excel_path_met)

        excel_path_stat = find_excel_workbook(XL_STAT)
        create_backup(excel_path_stat)

        combined_dfs_met = update_excel_workbook(hrv_metrics_dfs, excel_path_met, workbook_mode='metrics')
        _ = update_excel_workbook(hrv_metrics_dfs, excel_path_stat, workbook_mode='stats')

        ledger_int = IntervalLedger(str(INTERVALS_JSON))
        ledger_int.add_intervals(gaps)
        print(f"[INFO] ☑️ Interval ledger updated with {len(gaps)} interval gap(s)")

        print("[INFO] 🧹 Clearing checkpoints after successful Excel update.")
        for tf in TIMEFRAMES:
            if tf in TF:
                cp_path = os.path.join(checkpoint_dir, f"hrv_checkpoint_{tf}.parquet")
                if os.path.exists(cp_path): os.remove(cp_path)
        if os.path.exists(gaps_checkpoint_path): os.remove(gaps_checkpoint_path)

    except Exception as e:
        print(f"\n[ERROR] ❌ FAILED TO UPDATE EXCEL WORKBOOKS: {e}")
        print("[WARNING] ❗ Checkpoints have been saved. Re-run script without --fresh to try again from this point.")
        traceback.print_exc()

    duration = datetime.now() - start_time
    print(f"[INFO] 🎉 Standard HRV Analysis Completed in {duration}")

    if args.a:
        print("\n[INFO] 🧮 Calculating Multiscale Entropy & Correlaton Dimension...")
        run_entropic_metrics()

    if args.s:
        print("\n[INFO] === Calling Sequent Scripts ===")
        for script_num in args.s:
            pattern = os.path.join(SCRIPTS_DIR, f"HRV{script_num}*")
            matches = glob.glob(pattern)
            if matches:
                try:
                    subprocess.run(["python", matches[0]], check=True, cwd=SCRIPTS_DIR)
                    print(f"\n[INFO] 🎉 Finished {os.path.basename(matches[0])}")
                except Exception as e:
                    print(f"[ERROR] ❌ Failed to run {matches[0]}: {e}")
            else:
                print(f"[WARNING] ⚠️ No script found matching 'HRV{script_num}*' in {SCRIPTS_DIR}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[CRITICAL ERROR] ‼️ An unexpected error occurred in main execution: {e}")
        traceback.print_exc()